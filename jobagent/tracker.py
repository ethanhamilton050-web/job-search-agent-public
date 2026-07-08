"""Application tracker: Excel export of the SQLite pipeline."""
from __future__ import annotations

from pathlib import Path

from . import config, database


def export_xlsx(path: Path | None = None) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    path = path or (config.OUTPUT_DIR / "application_tracker.xlsx")
    path.parent.mkdir(parents=True, exist_ok=True)

    conn = database.connect()
    try:
        rows = database.ranked_listings(conn, min_score=0)
    finally:
        conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Applications"
    headers = ["Score", "Status", "Title", "Company", "Location",
               "Source", "URL", "Doc", "Follow-up", "Notes"]
    ws.append(headers)
    header_fill = PatternFill("solid", fgColor="305496")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill

    for r in rows:
        ws.append([
            r["score"], r["status"], r["title"], r["company"], r["location"],
            r["source"], r["url"], r["doc_path"] or "", r["follow_up"] or "",
            r["notes"] or "",
        ])

    widths = [8, 12, 30, 18, 18, 16, 40, 30, 12, 30]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"

    wb.save(path)
    return path
